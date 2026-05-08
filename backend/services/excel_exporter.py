"""
Phase 5 – Excel Exporter

Generates a professional multi-sheet .xlsx workbook from the financial model
and (optionally) forecast data using openpyxl.
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

from models import FinancialModelEntry, ForecastEntry, Project

logger = logging.getLogger(__name__)

# ── colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL = "1F3864"   # dark navy
SUBHEADER_FILL = "2E75B6"  # blue
DERIVED_FILL = "D6E4F0"   # light blue
TOTAL_FILL = "BDD7EE"    # mid blue
FORECAST_FILL = "E2EFDA"  # light green
HEADER_FONT_COLOR = "FFFFFF"


def _fmt_num(v: Optional[float]) -> Any:
    if v is None:
        return ""
    return round(v, 2)


def _write_statement_sheet(
    wb,
    sheet_name: str,
    periods: List[str],
    forecast_periods: List[str],
    rows: List[Dict[str, Any]],
    title: str,
):
    """Write one financial statement to a worksheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    ws = wb.create_sheet(title=sheet_name)
    all_periods = periods + forecast_periods

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_periods) + 1)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FONT_COLOR)
    ws["A1"].fill = hdr_fill(HEADER_FILL)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2: period headers
    ws["A2"] = "Line Item"
    ws["A2"].font = Font(bold=True, color=HEADER_FONT_COLOR)
    ws["A2"].fill = hdr_fill(HEADER_FILL)
    ws["A2"].alignment = Alignment(horizontal="left")

    for ci, period in enumerate(all_periods, start=2):
        cell = ws.cell(row=2, column=ci)
        cell.value = period
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.fill = hdr_fill(HEADER_FILL if period not in forecast_periods else SUBHEADER_FILL)
        cell.alignment = Alignment(horizontal="right")

    # Data rows
    for ri, row in enumerate(rows, start=3):
        label_cell = ws.cell(row=ri, column=1)
        label_cell.value = row["standard_label"]
        is_derived = row.get("is_derived", False)
        if is_derived:
            label_cell.font = Font(bold=True)
            label_cell.fill = hdr_fill(DERIVED_FILL)
        label_cell.alignment = Alignment(indent=1)
        label_cell.border = border

        for ci, period in enumerate(all_periods, start=2):
            val = row["values"].get(period)
            cell = ws.cell(row=ri, column=ci)
            cell.value = _fmt_num(val)
            cell.number_format = '#,##0.0'
            if period in forecast_periods:
                cell.fill = hdr_fill(FORECAST_FILL)
            elif is_derived:
                cell.fill = hdr_fill(DERIVED_FILL)
            cell.alignment = Alignment(horizontal="right")
            cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 38
    for ci in range(2, len(all_periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = "B3"


def _write_ratios_sheet(
    wb,
    ratios: Dict[str, Dict[str, Optional[float]]],
    periods: List[str],
):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed.")

    ws = wb.create_sheet(title="Key Ratios")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periods) + 1)
    ws["A1"] = "Key Financial Ratios"
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FONT_COLOR)
    ws["A1"].fill = hdr_fill(HEADER_FILL)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Ratio"
    ws["A2"].font = Font(bold=True, color=HEADER_FONT_COLOR)
    ws["A2"].fill = hdr_fill(HEADER_FILL)

    for ci, p in enumerate(periods, start=2):
        cell = ws.cell(row=2, column=ci)
        cell.value = p
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.fill = hdr_fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal="right")

    RATIO_LABELS = {
        "gross_margin": "Gross Margin (%)",
        "operating_margin": "Operating Margin (%)",
        "net_margin": "Net Margin (%)",
        "ebitda_margin": "EBITDA Margin (%)",
        "return_on_assets": "Return on Assets (%)",
        "return_on_equity": "Return on Equity (%)",
        "debt_to_equity": "Debt / Equity (x)",
        "fcf_margin": "FCF Margin (%)",
        "capex_to_revenue": "CapEx / Revenue (%)",
    }

    for ri, (key, label) in enumerate(RATIO_LABELS.items(), start=3):
        ws.cell(row=ri, column=1).value = label
        ws.cell(row=ri, column=1).border = border
        ws.cell(row=ri, column=1).alignment = Alignment(indent=1)
        for ci, p in enumerate(periods, start=2):
            val = ratios.get(key, {}).get(p)
            cell = ws.cell(row=ri, column=ci)
            cell.value = _fmt_num(val)
            cell.number_format = '0.0"%"' if "equity" not in key or "%" in label else '0.00x'
            cell.alignment = Alignment(horizontal="right")
            cell.border = border

    ws.column_dimensions["A"].width = 30
    for ci in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# ── public function ────────────────────────────────────────────────────────────

def export_to_excel(
    project_id: str,
    db: Session,
    include_forecast: bool = True,
) -> bytes:
    """
    Build the Excel workbook and return as bytes.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

    project = db.query(Project).filter(Project.id == project_id).first()
    company = project.company_name if project else "Company"

    # Load historical model entries
    hist_entries = (
        db.query(FinancialModelEntry)
        .filter(FinancialModelEntry.project_id == project_id)
        .all()
    )
    if not hist_entries:
        raise ValueError("No financial model data found. Build the model first.")

    # Organise by statement
    hist_periods_set = set()
    by_stmt: Dict[str, Dict[str, Dict[str, Optional[float]]]] = defaultdict(lambda: defaultdict(dict))
    derived_flags: Dict[str, bool] = {}
    for e in hist_entries:
        by_stmt[e.statement_type][e.standard_id][e.period] = e.value
        hist_periods_set.add(e.period)
        if e.is_derived:
            derived_flags[e.standard_id] = True

    hist_periods = sorted(hist_periods_set)

    # Load forecast entries
    forecast_periods: List[str] = []
    forecast_data: Dict[str, Dict[str, Optional[float]]] = defaultdict(dict)
    if include_forecast:
        fc_entries = (
            db.query(ForecastEntry)
            .filter(ForecastEntry.project_id == project_id)
            .all()
        )
        fc_periods_set = set()
        for e in fc_entries:
            forecast_data[e.standard_id][e.period] = e.value
            fc_periods_set.add(e.period)
        forecast_periods = sorted(fc_periods_set)

    wb = Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    from services.model_builder import IS_ITEMS, BS_ITEMS, CF_ITEMS

    STMT_CONFIG = [
        ("income_statement", "Income Statement", "Income Statement", IS_ITEMS),
        ("balance_sheet", "Balance Sheet", "Balance Sheet", BS_ITEMS),
        ("cash_flow", "Cash Flow", "Cash Flow Statement", CF_ITEMS),
    ]

    for stmt_key, sheet_name, title, item_ids in STMT_CONFIG:
        stmt_data = by_stmt.get(stmt_key, {})
        rows = []
        for sid in item_ids:
            if sid not in stmt_data and sid not in forecast_data:
                continue
            hist_vals = stmt_data.get(sid, {})
            fc_vals = forecast_data.get(sid, {})
            combined = {**{p: hist_vals.get(p) for p in hist_periods},
                        **{p: fc_vals.get(p) for p in forecast_periods}}
            rows.append({
                "standard_id": sid,
                "standard_label": sid.replace("_", " ").title(),
                "values": combined,
                "is_derived": derived_flags.get(sid, False),
            })
        if rows:
            _write_statement_sheet(
                wb, sheet_name, hist_periods, forecast_periods, rows,
                f"{company} — {title}"
            )

    # Ratios sheet (historical only for simplicity)
    # We'll reconstruct ratios inline from the data we have
    from services.model_builder import _compute_ratios
    flat_data: Dict[str, Dict[str, Optional[float]]] = {}
    for stmt_data in by_stmt.values():
        for sid, period_vals in stmt_data.items():
            flat_data[sid] = period_vals
    ratios = _compute_ratios(flat_data)
    if ratios:
        _write_ratios_sheet(wb, ratios, hist_periods)

    # Cover sheet
    ws_cover = wb.create_sheet(title="Cover", index=0)
    from openpyxl.styles import Font, PatternFill, Alignment
    ws_cover.merge_cells("A1:F1")
    ws_cover["A1"] = f"Financial Model — {company}"
    ws_cover["A1"].font = Font(bold=True, size=18, color=HEADER_FONT_COLOR)
    ws_cover["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 40

    ws_cover["A3"] = "Generated by Financial Model Generator"
    ws_cover["A3"].font = Font(italic=True, color="888888")
    ws_cover["A4"] = f"Ticker: {project.ticker or 'N/A'}"
    ws_cover["A5"] = f"Sector: {project.sector or 'N/A'}"
    ws_cover["A6"] = f"Currency: {project.base_currency or 'N/A'}"
    ws_cover["A7"] = f"Fiscal Year End: {project.fiscal_year_end or 'N/A'}"
    ws_cover["A9"] = "Contents:"
    ws_cover["A9"].font = Font(bold=True)
    for i, (_, sn, _, _) in enumerate(STMT_CONFIG, start=10):
        ws_cover[f"A{i}"] = f"  • {sn}"
    ws_cover[f"A{10 + len(STMT_CONFIG)}"] = "  • Key Ratios"
    ws_cover.column_dimensions["A"].width = 40

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
